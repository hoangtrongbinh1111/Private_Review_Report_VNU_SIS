"""
Flask Documentation:     http://flask.pocoo.org/docs/
Jinja2 Documentation:    http://jinja.pocoo.org/2/documentation/
Werkzeug Documentation:  http://werkzeug.pocoo.org/documentation/
This file creates your application.
"""

from app import app, db
from flask import render_template, request, redirect, url_for, flash, make_response, send_file, jsonify
import pandas as pd
import pyexcel
from app.forms import UserForm
from app.models import User
import uuid
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from io import BytesIO
import os
from werkzeug.utils import secure_filename
import random
import string
import json
import math
from flask_caching import Cache
# import sqlite3

# CONSTANT
ADMIN_TK = "wdWszs1JZ4D8NLu87AxDoJACikZAt1bpt6LrxXVgVi" # extract name
ADMIN_ROW_INDEX = 264499
FINISH_ROW_INDEX = 900 # to Xac nhan cua can bo thu truong don vi
app.config['UPLOAD_FOLDER'] = os.getcwd()
app.config['CACHE_TYPE'] = 'simple'
ALLOWED_EXTENSIONS = {'xlsx'}

cache = Cache(app)

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def excel_to_json():
    """Chuyển file Excel thành JSON và lưu cache"""
    file_path = os.path.join(app.config['UPLOAD_FOLDER'], 'data.xlsx')
    if not os.path.exists(file_path):
        return []

    df = pd.read_excel(file_path)
    df = df.iloc[:FINISH_ROW_INDEX]  # Giới hạn số dòng
    records = df.to_dict(orient='records')

    json_path = os.path.join(app.config['UPLOAD_FOLDER'], 'data.json')
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(records, f, ensure_ascii=False, indent=4)
    
    cache.set("excel_data", records)  # Lưu JSON vào cache RAM
    return records
###
# Routing for your application.
###
def format_currency(amount):
    try:
        """
        Format a number as currency in Vietnamese dong (VND).
        """
        # Convert the number to a string and reverse it
        reversed_amount = str(amount)[::-1]
        
        # Insert a dot (.) after every three digits
        formatted_amount = '.'.join(reversed_amount[i:i+3] for i in range(0, len(reversed_amount), 3))
        
        # Reverse the formatted string back to the original order
        formatted_amount = formatted_amount[::-1]
        
        # Add the currency symbol (₫)
        formatted_amount += ' ₫'
    except:
        formatted_amount = amount
    return formatted_amount

def generate_random_string(length=6):
    characters = string.ascii_letters + string.digits
    random_string = ''.join(random.choices(characters, k=length))
    return random_string

# Hàm kiểm tra NaN và thay thế bằng chuỗi rỗng
def clean_value(value):
    if value is None or value == '' or (isinstance(value, float) and math.isnan(value)):
        return ""  # Nếu là NaN, None hoặc chuỗi rỗng, thay bằng ""
    return value 

@app.route('/home')
def home():
    uuid = request.cookies.get('uuid')
    if not uuid:
        return redirect(url_for('login'))
    user = User.query.filter_by(uuid=uuid).first()
    print(user, user.uuid, ADMIN_TK)
    if uuid == ADMIN_TK:
        return render_template('admin.html')
    records = cache.get("excel_data")
    if records is None:
        records = excel_to_json()
    # Read the Excel file
    # records = pyexcel.get_records(file_name='data.xlsx')  # Replace 'data.xls' with the path to your Excel file
    # del records[0] # remove blank record
    # del records[FINISH_ROW_INDEX:]
    # Format the records with specific columns
    formatted_records = []
    for index, record in enumerate(records):
        data = list(record.values())
        # print(list(record.keys()))
        # print(list(record.values()))
        currency_money = data[user.rowIndex]
        if (
            currency_money is None or 
            currency_money == '' or 
            (isinstance(currency_money, (int, float)) and math.isnan(currency_money))
        ):
            continue
        try:
            if index >= 10:
                if int(currency_money) == 0:
                    continue
                if currency_money is not None and currency_money != '':
                    currency_money = format_currency(int(currency_money))
        except:
            currency_money = data[user.rowIndex]

        # Select specific columns from the record
        formatted_record = {
            'Ngày, tháng': clean_value(data[1]),
            # 'CK': clean_value(data[2]),
            'TM/CK': clean_value(data[3]),
            'Nội dung': clean_value(data[4]),
            'Số liệu': clean_value(currency_money),
            # Thêm các cột khác nếu cần
        }
        formatted_records.append(formatted_record)
    # Render the records in an HTML table
    return render_template('index.html', records=formatted_records, username=user.name)

@app.route('/', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        uuid = request.form['uuid']
        # Check if the UUID exists in the User table
        user = User.query.filter_by(uuid=uuid).first()
        if user:
            # Create a response object
            response = make_response(redirect(url_for('home')))
            # Set the cookie
            response.set_cookie('uuid', uuid, max_age=3600)

            # Redirect to the home page after successful login
            return response
        else:
            # Flash an alert message if login fails
            flash('UUID không chính xác. Xin mời nhập lại!', 'error')
            # Redirect back to the login page
            return redirect(url_for('login'))
    return render_template('login.html')

@app.route('/logout')
def logout():
    # Create a response object
    response = make_response(redirect(url_for('login')))
    # Clear the UUID cookie by setting its value to an empty string and setting its max_age to 0
    response.set_cookie('uuid', '', max_age=0)
    return response

@app.route('/upload', methods=['GET', 'POST'])
def upload_file():
    if 'file' not in request.files:
        return jsonify({'status': 'error', 'message': 'No file part'})
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'status': 'error', 'message': 'No selected file'})
    
    if file and allowed_file(file.filename):
        filename = 'data.xlsx'  # Fixed filename
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(file_path)

        # Xóa cache và cập nhật dữ liệu mới
        cache.delete("excel_data")
        excel_to_json()
        return jsonify({'status': 'success', 'message': 'File successfully uploaded'})
    else:
        return jsonify({'status': 'error', 'message': 'Allowed file types are xlsx only'})

@app.route('/clear_db')
def clear_db():
    try:
        db.create_all()
        # Delete all records from the User table
        User.query.delete()
        # Commit the changes to the database
        db.session.commit()
        return "All data in the User table deleted successfully."
    except Exception as e:
        # If any exception occurs, print the error message
        print(f"An error occurred: {str(e)}")
        # Optionally, you can rollback the database session in case of an error
        db.session.rollback()
        return "Error"
    
@app.route('/import_excel')
def import_excel():
    # Check if the user already exists
    existing_user = User.query.filter_by(name="ADMIN").first()
    if existing_user is None:
        # User does not exist, create and add to the session
        user = User("9999", "ADMIN", str(ADMIN_TK), ADMIN_ROW_INDEX)
        db.session.add(user)
        db.session.commit()

    status = True
    try:
        records = cache.get("excel_data")
        if records is None:
            records = excel_to_json()
        # Read the Excel file
        # records = pyexcel.get_records(file_name='data.xlsx')  # Replace 'data.xls' with the path to your Excel file
        list_user = list(records[0].keys()) # danh sách users
        # Danh sách mã số
        list_code = list(records[0].values())
        start_user_index = 5 # bắt đầu từ index User
        end_user_index = len(list_user) # 
        list_user_format = list_user[start_user_index:end_user_index]
        for user_name in list_user_format:
            if clean_value(user_name) == "":
                continue
            # Check if the user already exists
            existing_user = User.query.filter_by(code=str(list_code[start_user_index])).first()
            # If the user does not exist, create and add to the session
            if existing_user is None:
                # User does not exist, create and add to the session
                user = User(str(list_code[start_user_index]), user_name, generate_random_string(), start_user_index)
                db.session.add(user)
            start_user_index = start_user_index + 1
        db.session.commit()
    except Exception as e:
        # If any exception occurs, print the error message
        print(f"An error occurred: {str(e)}")
        # Optionally, you can rollback the database session in case of an error
        db.session.rollback()
        status = False
    return render_template('about.html', status=status)

def generate_message(name, uuid):
    message = f"Phòng KHTC xin gửi thầy, cô đường link https://taichinh.sis.vnu.edu.vn và mật khẩu {uuid} để truy cập dữ liệu thu nhập hàng tháng. Trân trọng!"
    return message
@app.route('/export_excel')
def export_excel():
    uuid = request.cookies.get('uuid')
    print(uuid, ADMIN_TK, uuid== ADMIN_TK)
    if not uuid or uuid != ADMIN_TK:
        return redirect(url_for('login'))
    # Create a new Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.append(["UUID", "Họ và tên", "Lời nhắn"])
    # Set column widths and styles for header row
    header_row = ws[1]
    for cell in header_row:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        column_letter = get_column_letter(cell.column)
        ws.column_dimensions[column_letter].width = 50  # Set width for all columns

    # Add data to the worksheet
    users = db.session.query(User).all()
    for row_index, user in enumerate(users, start=2):  # Start from row 2 for data rows
        ws.cell(row=row_index, column=1, value=user.uuid)
        ws.cell(row=row_index, column=2, value=user.name)
        message = generate_message(user.name, user.uuid)
        ws.cell(row=row_index, column=3, value=message)
        # Adjust column width to fit content
        # Adjust column width to fit content and set alignment to justified
    for col in range(1, 4):  # Adjusting columns 1 to 3
        column_letter = get_column_letter(col)
        ws.column_dimensions[column_letter].auto_size = True  # Auto size column width
        for row in ws.iter_rows(min_row=2, min_col=col, max_row=ws.max_row, max_col=col):
            for cell in row:
                cell.alignment = Alignment(horizontal='justify')  # Justify align for data cells

    # Save the workbook to a BytesIO buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Send the file to the client for download
    return send_file(buffer, as_attachment=True, attachment_filename='data.xlsx')

@app.route('/users')
def show_users():
    # uuid = request.cookies.get('uuid')
    # if not uuid or uuid != str(ADMIN_TK):
    #     return redirect(url_for('login'))
    users = db.session.query(User).all() # or you could have used User.query.all()

    return render_template('show_users.html', users=users)

# Flash errors from the form if validation fails
def flash_errors(form):
    for field, errors in form.errors.items():
        for error in errors:
            flash(u"Error in the %s field - %s" % (
                getattr(form, field).label.text,
                error
            ))

###
# The functions below should be applicable to all Flask apps.
###

@app.route('/<file_name>.txt')
def send_text_file(file_name):
    """Send your static text file."""
    file_dot_text = file_name + '.txt'
    return app.send_static_file(file_dot_text)


@app.after_request
def add_header(response):
    """
    Add headers to both force latest IE rendering engine or Chrome Frame,
    and also to cache the rendered page for 10 minutes.
    """
    response.headers['X-UA-Compatible'] = 'IE=Edge,chrome=1'
    response.headers['Cache-Control'] = 'public, max-age=600'
    return response


@app.errorhandler(404)
def page_not_found(error):
    """Custom 404 page."""
    return render_template('404.html'), 404


if __name__ == '__main__':
    app.run(debug=True,host="0.0.0.0",port="8080")
